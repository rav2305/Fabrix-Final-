import os
import urllib.parse
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
import csv
import io

from models import db, User, Product, Invoice, InvoiceItem, DealerPurchase

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'fabrix-secret-key-12345')

# Support dynamic Database URI (SQLAlchemy PostgreSQL / SQLite fallback)
database_url = os.environ.get('DATABASE_URL')
if database_url:
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///fabrix.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# Login configuration
login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.login_message_category = 'error'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Create database tables and seed default admin on startup
with app.app_context():
    db.create_all()
    # Seed default admin if no users exist
    if User.query.count() == 0:
        admin_user = User(
            username='admin',
            password_hash=generate_password_hash('admin123'),
            role='admin'
        )
        db.session.add(admin_user)
        db.session.commit()
        print("--- DEFAULT ADMIN SEEDED: admin / admin123 ---")

# Context processor to inject variables globally if needed
@app.context_processor
def inject_now():
    return {'now': datetime.utcnow()}

# ----------------- AUTH ROUTES -----------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            flash("Welcome back to Fabrix!", "success")
            return redirect(url_for('dashboard'))
        else:
            flash("Invalid username or password.", "error")
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash("You have logged out successfully.", "success")
    return redirect(url_for('login'))

# ----------------- DASHBOARD ROUTE -----------------

@app.route('/')
@login_required
def dashboard():
    # Gather reports metrics
    invoices = Invoice.query.all()
    sales_revenue = sum(inv.final_amount for inv in invoices)
    total_invoices_count = len(invoices)
    
    # Calculate sales profit margin
    net_sales_profit = 0.0
    for inv in invoices:
        for item in inv.items:
            net_sales_profit += (item.selling_price - item.cost_price) * item.quantity
        # Subtract discounts
        net_sales_profit -= inv.discount
        
    # Gather inventory values
    products = Product.query.all()
    stock_value_cost = sum(prod.quantity * prod.cost_price for prod in products)
    total_stock_qty = sum(prod.quantity for prod in products)
    
    # Gather dealer outstanding liabilities
    purchases = DealerPurchase.query.all()
    dealer_outstanding = sum(p.outstanding_amount for p in purchases)
    
    # Recent Invoices (limit to last 5)
    recent_invoices = Invoice.query.order_by(Invoice.date_created.desc()).limit(5).all()
    
    # Low stock items (qty <= 10)
    low_stock_items = Product.query.filter(Product.quantity <= 10).order_by(Product.quantity.asc()).all()
    
    return render_template(
        'dashboard.html',
        sales_revenue=sales_revenue,
        total_invoices_count=total_invoices_count,
        net_sales_profit=net_sales_profit,
        stock_value_cost=stock_value_cost,
        total_stock_qty=total_stock_qty,
        dealer_outstanding=dealer_outstanding,
        recent_invoices=recent_invoices,
        low_stock_items=low_stock_items
    )

# ----------------- STOCK MANAGEMENT ROUTES -----------------

@app.route('/stock')
@login_required
def stock():
    products = Product.query.order_by(Product.name.asc()).all()
    return render_template('stock.html', products=products)

@app.route('/stock/add', methods=['POST'])
@login_required
def add_product():
    name = request.form.get('name', '').strip()
    quantity = int(request.form.get('quantity', 0))
    selling_price = float(request.form.get('selling_price', 0.0))
    cost_price = float(request.form.get('cost_price', 0.0))
    
    if not name:
        flash("Product name cannot be empty.", "error")
        return redirect(url_for('stock'))
        
    existing = Product.query.filter_by(name=name).first()
    if existing:
        flash(f"Product '{name}' already exists. Use Edit to update stock.", "error")
        return redirect(url_for('stock'))
        
    prod = Product(name=name, quantity=quantity, selling_price=selling_price, cost_price=cost_price)
    db.session.add(prod)
    db.session.commit()
    
    flash(f"Product '{name}' added successfully.", "success")
    return redirect(url_for('stock'))

@app.route('/stock/edit', methods=['POST'])
@login_required
def edit_product():
    prod_id = request.form.get('id')
    name = request.form.get('name', '').strip()
    quantity = int(request.form.get('quantity', 0))
    selling_price = float(request.form.get('selling_price', 0.0))
    cost_price = float(request.form.get('cost_price', 0.0))
    
    prod = Product.query.get(prod_id)
    if not prod:
        flash("Product not found.", "error")
        return redirect(url_for('stock'))
        
    # Check unique name constraint if name is changed
    if name != prod.name:
        existing = Product.query.filter_by(name=name).first()
        if existing:
            flash(f"A product named '{name}' already exists.", "error")
            return redirect(url_for('stock'))
            
    prod.name = name
    prod.quantity = quantity
    prod.selling_price = selling_price
    prod.cost_price = cost_price
    db.session.commit()
    
    flash(f"Product '{name}' updated successfully.", "success")
    return redirect(url_for('stock'))

@app.route('/stock/delete/<int:product_id>', methods=['POST'])
@login_required
def delete_product(product_id):
    if not current_user.is_admin():
        flash("Only Administrators can delete product records.", "error")
        return redirect(url_for('stock'))
        
    prod = Product.query.get(product_id)
    if prod:
        db.session.delete(prod)
        db.session.commit()
        flash(f"Product deleted successfully.", "success")
    else:
        flash("Product not found.", "error")
    return redirect(url_for('stock'))

@app.route('/stock/download-template')
@login_required
def download_stock_template():
    import io
    from openpyxl import Workbook
    from flask import send_file
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Fabrix Stock Template"
    
    # Write headers
    headers = ["Product Name", "Qty", "Selling Price", "Price for us"]
    ws.append(headers)
    
    # Write some sample rows
    ws.append(["Sample Denim Jacket", 50, 1500.0, 900.0])
    ws.append(["Sample Cotton T-Shirt", 120, 450.0, 200.0])
    ws.append(["Sample Linen Trousers", 30, 1200.0, 700.0])
    
    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return send_file(
        file_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="fabrix_stock_template.xlsx"
    )

@app.route('/stock/bulk-upload', methods=['POST'])
@login_required
def bulk_upload():
    import json
    if 'excel_file' not in request.files:
        flash("No file uploaded.", "error")
        return redirect(url_for('stock'))
        
    file = request.files['excel_file']
    if file.filename == '':
        flash("No selected file.", "error")
        return redirect(url_for('stock'))

    try:
        rows_data = []
        if file.filename.endswith('.csv'):
            # Parse CSV
            file_content = file.read().decode('utf-8', errors='ignore')
            csv_data = io.StringIO(file_content)
            reader = csv.reader(csv_data)
            try:
                headers = [col.strip() for col in next(reader)]
            except StopIteration:
                flash("Uploaded CSV file is empty.", "error")
                return redirect(url_for('stock'))
                
            # Column mapping for fuzzy matching
            col_mapping = {}
            for i, col in enumerate(headers):
                col_lower = col.lower()
                if 'product name' in col_lower or 'product' in col_lower or 'name' in col_lower:
                    col_mapping['name'] = i
                elif 'qty' in col_lower or 'quantity' in col_lower:
                    col_mapping['quantity'] = i
                elif 'selling price' in col_lower or 'selling' in col_lower:
                    col_mapping['selling_price'] = i
                elif 'price for us' in col_lower or 'cost price' in col_lower or 'cost' in col_lower or 'us' in col_lower:
                    col_mapping['cost_price'] = i
            
            # Check required
            required = ['name', 'quantity', 'selling_price', 'cost_price']
            for req in required:
                if req not in col_mapping:
                    if len(headers) >= 4:
                        col_mapping = {'name': 0, 'quantity': 1, 'selling_price': 2, 'cost_price': 3}
                        break
                    else:
                        flash(f"CSV is missing required headers. Found headers: {headers}", "error")
                        return redirect(url_for('stock'))
            
            for row in reader:
                if not row or len(row) <= max(col_mapping.values()):
                    continue
                rows_data.append({
                    'name': row[col_mapping['name']],
                    'quantity': row[col_mapping['quantity']],
                    'selling_price': row[col_mapping['selling_price']],
                    'cost_price': row[col_mapping['cost_price']]
                })
        else:
            # Parse Excel using openpyxl
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            
            # Read first row for headers
            first_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
            if not first_row:
                flash("Uploaded Excel file is empty.", "error")
                return redirect(url_for('stock'))
                
            headers = [str(col).strip() if col is not None else "" for col in first_row]
            
            # Column mapping for fuzzy matching
            col_mapping = {}
            for i, col in enumerate(headers):
                col_lower = col.lower()
                if 'product name' in col_lower or 'product' in col_lower or 'name' in col_lower:
                    col_mapping['name'] = i
                elif 'qty' in col_lower or 'quantity' in col_lower:
                    col_mapping['quantity'] = i
                elif 'selling price' in col_lower or 'selling' in col_lower:
                    col_mapping['selling_price'] = i
                elif 'price for us' in col_lower or 'cost price' in col_lower or 'cost' in col_lower or 'us' in col_lower:
                    col_mapping['cost_price'] = i
            
            # Check required
            required = ['name', 'quantity', 'selling_price', 'cost_price']
            for req in required:
                if req not in col_mapping:
                    if len(headers) >= 4:
                        col_mapping = {'name': 0, 'quantity': 1, 'selling_price': 2, 'cost_price': 3}
                        break
                    else:
                        flash(f"Excel is missing required headers. Found headers: {headers}", "error")
                        return redirect(url_for('stock'))
            
            # Read rows starting from row 2
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(col_mapping.values()):
                    continue
                name_val = row[col_mapping['name']]
                rows_data.append({
                    'name': name_val,
                    'quantity': row[col_mapping['quantity']],
                    'selling_price': row[col_mapping['selling_price']],
                    'cost_price': row[col_mapping['cost_price']]
                })
        
        # Analyze and validate data
        valid_rows = []
        invalid_rows = []
        new_count = 0
        update_count = 0
        
        for index, row in enumerate(rows_data):
            name = row['name']
            qty_raw = row['quantity']
            selling_raw = row['selling_price']
            cost_raw = row['cost_price']
            row_num = index + 2
            
            if name is None or str(name).strip() == "" or str(name).lower() == 'nan':
                invalid_rows.append({
                    'row': row_num,
                    'product_name': '[Empty]',
                    'error': "Product Name is empty or invalid"
                })
                continue
                
            name_str = str(name).strip()
            
            # Parse qty
            try:
                if qty_raw is None or str(qty_raw).strip() == "":
                    qty = 0
                else:
                    qty = int(float(str(qty_raw).strip()))
                if qty < 0:
                    raise ValueError("Negative stock quantity")
            except Exception as e:
                invalid_rows.append({
                    'row': row_num,
                    'product_name': name_str,
                    'error': f"Invalid Quantity: '{qty_raw}' ({str(e)})"
                })
                continue
                
            # Parse selling price
            try:
                if selling_raw is None or str(selling_raw).strip() == "":
                    selling = 0.0
                else:
                    selling = float(str(selling_raw).strip().replace(',', ''))
                if selling < 0:
                    raise ValueError("Negative selling price")
            except Exception as e:
                invalid_rows.append({
                    'row': row_num,
                    'product_name': name_str,
                    'error': f"Invalid Selling Price: '{selling_raw}' ({str(e)})"
                })
                continue
                
            # Parse cost price
            try:
                if cost_raw is None or str(cost_raw).strip() == "":
                    cost = 0.0
                else:
                    cost = float(str(cost_raw).strip().replace(',', ''))
                if cost < 0:
                    raise ValueError("Negative cost price")
            except Exception as e:
                invalid_rows.append({
                    'row': row_num,
                    'product_name': name_str,
                    'error': f"Invalid Cost Price: '{cost_raw}' ({str(e)})"
                })
                continue
                
            prod_exists = Product.query.filter_by(name=name_str).first()
            if prod_exists:
                update_count += 1
                row_type = "Update"
            else:
                new_count += 1
                row_type = "New Product"
                
            valid_rows.append({
                'name': name_str,
                'quantity': qty,
                'selling_price': selling,
                'cost_price': cost,
                'type': row_type
            })
            
        # Write valid_rows to a temporary JSON file to handle large datasets safely
        temp_dir = os.path.join(app.instance_path, 'temp_imports')
        os.makedirs(temp_dir, exist_ok=True)
        temp_filepath = os.path.join(temp_dir, f"import_{current_user.id}.json")
        with open(temp_filepath, 'w') as f:
            json.dump(valid_rows, f)
            
        return render_template(
            'import_preview.html',
            total_count=len(rows_data),
            valid_count=len(valid_rows),
            invalid_count=len(invalid_rows),
            new_count=new_count,
            update_count=update_count,
            valid_rows=valid_rows,
            invalid_rows=invalid_rows
        )
    except Exception as e:
        flash(f"Bulk upload analysis failed: {str(e)}", "error")
        return redirect(url_for('stock'))

@app.route('/stock/bulk-upload/confirm', methods=['POST'])
@login_required
def confirm_bulk_upload():
    import json
    temp_filepath = os.path.join(app.instance_path, 'temp_imports', f"import_{current_user.id}.json")
    if not os.path.exists(temp_filepath):
        flash("No upload session found. Please re-upload your file.", "error")
        return redirect(url_for('stock'))
        
    try:
        with open(temp_filepath, 'r') as f:
            valid_rows = json.load(f)
            
        added = 0
        updated = 0
        for row in valid_rows:
            name = row['name']
            qty = row['quantity']
            selling = row['selling_price']
            cost = row['cost_price']
            
            prod = Product.query.filter_by(name=name).first()
            if prod:
                prod.quantity += qty
                prod.selling_price = selling
                prod.cost_price = cost
                updated += 1
            else:
                new_prod = Product(name=name, quantity=qty, selling_price=selling, cost_price=cost)
                db.session.add(new_prod)
                added += 1
                
        db.session.commit()
        os.remove(temp_filepath)
        flash(f"Success! Imported {added} new products and restocked {updated} existing products.", "success")
    except Exception as e:
        flash(f"Failed to write stock updates to database: {str(e)}", "error")
        
    return redirect(url_for('stock'))

@app.route('/stock/bulk-upload/cancel', methods=['POST'])
@login_required
def cancel_bulk_upload():
    temp_filepath = os.path.join(app.instance_path, 'temp_imports', f"import_{current_user.id}.json")
    if os.path.exists(temp_filepath):
        try:
            os.remove(temp_filepath)
        except:
            pass
    flash("Import action canceled.", "info")
    return redirect(url_for('stock'))

# ----------------- BILLING & INVOICE ROUTES -----------------

@app.route('/invoice/new', methods=['GET', 'POST'])
@login_required
def create_invoice():
    if request.method == 'POST':
        customer_name = request.form.get('customer_name', '').strip() or 'Walk-in Customer'
        customer_phone = request.form.get('customer_phone', '').strip()
        payment_method = request.form.get('payment_method', 'Cash')
        discount = float(request.form.get('discount', 0.0))
        
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('quantity[]')
        selling_prices = request.form.getlist('selling_price[]')
        
        if not product_ids:
            flash("Cannot create an empty invoice.", "error")
            return redirect(url_for('create_invoice'))
            
        # Generate Invoice Number
        date_str = datetime.now().strftime('%Y%m%d')
        # Count number of invoices today to generate sequential ID
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_count = Invoice.query.filter(Invoice.date_created >= today_start).count()
        invoice_number = f"ORD-{date_str}-{str(today_count + 1).zfill(3)}"
        
        # Calculate pricing
        subtotal = 0.0
        invoice_items = []
        
        for pid, qty_str, price_str in zip(product_ids, quantities, selling_prices):
            qty = int(qty_str)
            selling_price = float(price_str)
            prod = Product.query.get(pid)
            
            if prod:
                if prod.quantity < qty:
                    flash(f"Insufficient stock for '{prod.name}'. Available: {prod.quantity}, Requested: {qty}", "error")
                    return redirect(url_for('create_invoice'))
                
                # Reduce stock
                prod.quantity -= qty
                subtotal += selling_price * qty
                
                item = InvoiceItem(
                    product_name=prod.name,
                    quantity=qty,
                    selling_price=selling_price,
                    cost_price=prod.cost_price
                )
                invoice_items.append(item)
        
        final_amount = max(0.0, subtotal - discount)
        
        # Capture paid amount and calculate status
        amount_paid_raw = request.form.get('amount_paid', '').strip()
        if amount_paid_raw == '':
            amount_paid = final_amount
        else:
            amount_paid = float(amount_paid_raw)
            
        if amount_paid >= final_amount:
            amount_paid = final_amount
            payment_status = 'Paid'
        elif amount_paid > 0:
            payment_status = 'Partial'
        else:
            payment_status = 'Unpaid'
            
        # Save invoice record
        invoice = Invoice(
            invoice_number=invoice_number,
            customer_name=customer_name,
            customer_phone=customer_phone,
            discount=discount,
            total_amount=subtotal,
            final_amount=final_amount,
            amount_paid=amount_paid,
            payment_status=payment_status,
            payment_method=payment_method,
            items=invoice_items
        )
        
        db.session.add(invoice)
        db.session.commit()
        
        flash(f"Invoice {invoice_number} created successfully.", "success")
        return redirect(url_for('invoice_detail', invoice_id=invoice.id))
        
    products = Product.query.order_by(Product.name.asc()).all()
    return render_template('invoice_create.html', products=products)

@app.route('/invoice/<int:invoice_id>')
@login_required
def invoice_detail(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    
    # Generate WhatsApp message body based on exact user requested template
    customer_name = invoice.customer_name or 'PREET'
    total_display = f"{int(invoice.final_amount)}" if invoice.final_amount.is_integer() else f"{invoice.final_amount:,.2f}"
    
    # Generate public viewing URL to download PDF
    public_url = f"{request.url_root}invoice/public/{invoice.invoice_number}"
    
    msg_template = (
        f"Hello {customer_name},\n\n"
        f"Thanks for visiting Fabrix and shopping with us! 🛍️\n\n"
        f"Join our community to stay updated with new collections and deals:  https://chat.whatsapp.com/DOkCF1xxaQeB0kjzjLu8UR\n\n"
        f"Your invoice for order #{invoice.invoice_number} is attached here.\n"
        f"Total: ₹{total_display}\n"
        f"View & Download PDF Invoice: {public_url}\n\n"
        f"📍 SUPER MALL-2, FF/152, Infocity, Gandhinagar, Gujarat 382007\n\n"
        f"Thanks again for your purchase—hope to see you again soon! 💫"
    )
    
    # URL encode parameters
    encoded_message = urllib.parse.quote(msg_template)
    if invoice.customer_phone:
        # Prepend country code if missing (assumes Indian standard +91)
        phone = invoice.customer_phone
        if len(phone) == 10:
            phone = "91" + phone
        whatsapp_url = f"https://api.whatsapp.com/send?phone={phone}&text={encoded_message}"
    else:
        whatsapp_url = f"https://api.whatsapp.com/send?text={encoded_message}"
        
    return render_template('invoice_detail.html', invoice=invoice, whatsapp_url=whatsapp_url)

@app.route('/invoice/public/<string:invoice_number>')
def public_invoice(invoice_number):
    invoice = Invoice.query.filter_by(invoice_number=invoice_number).first_or_404()
    return render_template('invoice_detail.html', invoice=invoice, is_public=True)

@app.route('/invoices')
@login_required
def invoices_history():
    invoice_number = request.args.get('invoice_number', '').strip()
    customer_name = request.args.get('customer_name', '').strip()
    payment_status = request.args.get('payment_status', '').strip()
    start_date_str = request.args.get('start_date', '').strip()
    end_date_str = request.args.get('end_date', '').strip()
    
    query = Invoice.query
    
    if invoice_number:
        query = query.filter(Invoice.invoice_number.ilike(f"%{invoice_number}%"))
        
    if customer_name:
        query = query.filter(Invoice.customer_name.ilike(f"%{customer_name}%"))
        
    if payment_status:
        query = query.filter(Invoice.payment_status == payment_status)
        
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(Invoice.date_created >= start_date)
        except ValueError:
            pass
            
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(Invoice.date_created <= end_date)
        except ValueError:
            pass
            
    invoices = query.order_by(Invoice.date_created.desc()).all()
    return render_template('invoices.html', invoices=invoices)

@app.route('/invoice/update-payment', methods=['POST'])
@login_required
def update_invoice_payment():
    invoice_id = request.form.get('id')
    amount_received = float(request.form.get('amount_received') or 0.0)
    payment_method = request.form.get('payment_method', 'Cash')
    
    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        flash("Invoice not found.", "error")
        return redirect(url_for('dashboard'))
        
    if amount_received <= 0:
        flash("Please enter a valid payment amount.", "error")
        return redirect(url_for('invoice_detail', invoice_id=invoice.id))
        
    if amount_received > invoice.outstanding_amount:
        flash(f"Amount received (₹{amount_received}) cannot exceed outstanding balance (₹{invoice.outstanding_amount}).", "error")
        return redirect(url_for('invoice_detail', invoice_id=invoice.id))
        
    invoice.amount_paid += amount_received
    invoice.payment_method = payment_method
    
    # Recalculate status
    if invoice.amount_paid >= invoice.final_amount:
        invoice.amount_paid = invoice.final_amount
        invoice.payment_status = 'Paid'
    elif invoice.amount_paid > 0:
        invoice.payment_status = 'Partial'
    else:
        invoice.payment_status = 'Unpaid'
        
    db.session.commit()
    flash(f"Payment of ₹{amount_received:.2f} recorded successfully.", "success")
    return redirect(url_for('invoice_detail', invoice_id=invoice.id))

# ----------------- DEALER PURCHASES ROUTES -----------------

@app.route('/purchases')
@login_required
def purchases():
    purchases = DealerPurchase.query.order_by(DealerPurchase.date_created.desc()).all()
    total_purchase_cost = sum(p.total_cost for p in purchases)
    total_purchase_paid = sum(p.amount_paid for p in purchases)
    
    return render_template(
        'purchases.html',
        purchases=purchases,
        total_purchase_cost=total_purchase_cost,
        total_purchase_paid=total_purchase_paid
    )

@app.route('/purchases/add', methods=['POST'])
@login_required
def add_purchase():
    dealer_name = request.form.get('dealer_name', '').strip()
    product_description = request.form.get('product_description', '').strip()
    quantity = int(request.form.get('quantity') or 0)
    total_cost = float(request.form.get('total_cost') or 0.0)
    amount_paid = float(request.form.get('amount_paid') or 0.0)
    
    if not dealer_name:
        flash("Dealer name is required.", "error")
        return redirect(url_for('purchases'))
        
    # Determine status
    if amount_paid >= total_cost:
        status = 'Paid'
    elif amount_paid > 0:
        status = 'Partial'
    else:
        status = 'Unpaid'
        
    purchase = DealerPurchase(
        dealer_name=dealer_name,
        product_description=product_description,
        quantity=quantity,
        total_cost=total_cost,
        amount_paid=amount_paid,
        payment_status=status
    )
    db.session.add(purchase)
    db.session.commit()
    
    flash(f"Purchase from '{dealer_name}' logged successfully.", "success")
    return redirect(url_for('purchases'))

@app.route('/purchases/update-payment', methods=['POST'])
@login_required
def update_purchase_payment():
    purchase_id = request.form.get('id')
    amount_paid = float(request.form.get('amount_paid') or 0.0)
    
    purchase = DealerPurchase.query.get(purchase_id)
    if not purchase:
        flash("Purchase record not found.", "error")
        return redirect(url_for('purchases'))
        
    purchase.amount_paid = amount_paid
    
    # Update status
    if purchase.amount_paid >= purchase.total_cost:
        purchase.payment_status = 'Paid'
    elif purchase.amount_paid > 0:
        purchase.payment_status = 'Partial'
    else:
        purchase.payment_status = 'Unpaid'
        
    db.session.commit()
    flash(f"Payment history for {purchase.dealer_name} updated.", "success")
    return redirect(url_for('purchases'))

@app.route('/purchases/delete/<int:purchase_id>', methods=['POST'])
@login_required
def delete_purchase(purchase_id):
    if not current_user.is_admin():
        flash("Only Administrators can delete purchase ledgers.", "error")
        return redirect(url_for('purchases'))
        
    purchase = DealerPurchase.query.get(purchase_id)
    if purchase:
        db.session.delete(purchase)
        db.session.commit()
        flash("Purchase ledger deleted.", "success")
    else:
        flash("Purchase ledger not found.", "error")
    return redirect(url_for('purchases'))

# ----------------- REPORTS ROUTE -----------------

@app.route('/reports')
@login_required
def reports():
    invoices = Invoice.query.order_by(Invoice.date_created.desc()).all()
    products = Product.query.order_by(Product.name.asc()).all()
    purchases = DealerPurchase.query.order_by(DealerPurchase.date_created.desc()).all()
    
    # Telemetry metrics
    sales_revenue = sum(inv.final_amount for inv in invoices)
    total_invoices_count = len(invoices)
    
    # Sales Profit Margin calculation
    net_sales_profit = 0.0
    for inv in invoices:
        for item in inv.items:
            net_sales_profit += (item.selling_price - item.cost_price) * item.quantity
        net_sales_profit -= inv.discount
        
    # Stock counts
    total_stock_qty = sum(prod.quantity for prod in products)
    stock_value_cost = sum(prod.quantity * prod.cost_price for prod in products)
    stock_value_retail = sum(prod.quantity * prod.selling_price for prod in products)
    
    # Dealer accounts
    dealer_total_bill = sum(p.total_cost for p in purchases)
    dealer_total_paid = sum(p.amount_paid for p in purchases)
    
    return render_template(
        'reports.html',
        invoices=invoices,
        products=products,
        purchases=purchases,
        sales_revenue=sales_revenue,
        total_invoices_count=total_invoices_count,
        net_sales_profit=net_sales_profit,
        total_stock_qty=total_stock_qty,
        stock_value_cost=stock_value_cost,
        stock_value_retail=stock_value_retail,
        dealer_total_bill=dealer_total_bill,
        dealer_total_paid=dealer_total_paid
    )

# ----------------- USER ACCESS MANAGEMENT ROUTES -----------------

@app.route('/users')
@login_required
def users():
    if not current_user.is_admin():
        flash("Unauthorized access: Administrators only.", "error")
        return redirect(url_for('dashboard'))
        
    users = User.query.order_by(User.id.asc()).all()
    return render_template('users.html', users=users)

@app.route('/users/add', methods=['POST'])
@login_required
def add_user():
    if not current_user.is_admin():
        flash("Unauthorized action.", "error")
        return redirect(url_for('dashboard'))
        
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    role = request.form.get('role', 'staff')
    
    if not username or not password:
        flash("Username and password cannot be empty.", "error")
        return redirect(url_for('users'))
        
    existing = User.query.filter_by(username=username).first()
    if existing:
        flash(f"Username '{username}' is already taken.", "error")
        return redirect(url_for('users'))
        
    user = User(
        username=username,
        password_hash=generate_password_hash(password),
        role=role
    )
    db.session.add(user)
    db.session.commit()
    
    flash(f"User '{username}' created successfully.", "success")
    return redirect(url_for('users'))

@app.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_admin():
        flash("Unauthorized action.", "error")
        return redirect(url_for('dashboard'))
        
    if user_id == current_user.id:
        flash("You cannot delete your own active administrator account.", "error")
        return redirect(url_for('users'))
        
    user = User.query.get(user_id)
    if user:
        db.session.delete(user)
        db.session.commit()
        flash(f"User access revoked successfully.", "success")
    else:
        flash("User not found.", "error")
    return redirect(url_for('users'))

@app.route('/admin/backup-db')
@login_required
def backup_database():
    if not current_user.is_admin():
        flash("Unauthorized access.", "error")
        return redirect(url_for('dashboard'))
        
    db_path = os.path.join(app.instance_path, 'fabrix.db')
    if os.path.exists(db_path) and app.config['SQLALCHEMY_DATABASE_URI'].startswith('sqlite'):
        from flask import send_file
        return send_file(
            db_path,
            as_attachment=True,
            download_name=f"fabrix_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        )
    else:
        import json
        from flask import send_file
        data = {
            'users': [{'username': u.username, 'role': u.role} for u in User.query.all()],
            'products': [{'name': p.name, 'quantity': p.quantity, 'selling_price': p.selling_price, 'cost_price': p.cost_price} for p in Product.query.all()],
            'invoices': [{
                'invoice_number': i.invoice_number,
                'customer_name': i.customer_name,
                'customer_phone': i.customer_phone,
                'total_amount': i.total_amount,
                'discount': i.discount,
                'final_amount': i.final_amount,
                'amount_paid': i.amount_paid,
                'payment_status': i.payment_status,
                'payment_method': i.payment_method,
                'date_created': i.date_created.isoformat(),
                'items': [{'product_name': item.product_name, 'quantity': item.quantity, 'selling_price': item.selling_price, 'cost_price': item.cost_price} for item in i.items]
            } for i in Invoice.query.all()],
            'purchases': [{
                'dealer_name': p.dealer_name,
                'product_description': p.product_description,
                'quantity': p.quantity,
                'total_cost': p.total_cost,
                'amount_paid': p.amount_paid,
                'payment_status': p.payment_status,
                'date_created': p.date_created.isoformat()
            } for p in DealerPurchase.query.all()]
        }
        
        file_stream = io.BytesIO()
        file_stream.write(json.dumps(data, indent=4).encode('utf-8'))
        file_stream.seek(0)
        
        return send_file(
            file_stream,
            mimetype="application/json",
            as_attachment=True,
            download_name=f"fabrix_db_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )

# Run server

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', use_reloader=False)
